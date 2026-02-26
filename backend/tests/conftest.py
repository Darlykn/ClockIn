"""
conftest.py — shared fixtures for all QA tests.

Strategy:
- Use the same PostgreSQL as the running Docker app (DATABASE_URL from settings).
- Each fixture-created user has a unique username prefixed with "qa_" to avoid
  collisions with seed data and between test runs.
- Session-scoped fixtures (admin_user, admin_token) create data once and clean up
  after the full test session.
- Function-scoped fixtures (employee_user, manager_user, fresh_user) create data
  per test and clean up afterwards.
"""

from __future__ import annotations

import io
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pyotp
import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.security import hash_password
from app.db.models import AttendanceLog, ImportHistory, User
from app.main import app

# ---------------------------------------------------------------------------
# Database engine for test fixtures (independent of app's get_db)
# ---------------------------------------------------------------------------
_test_engine = create_async_engine(settings.DATABASE_URL, echo=False, pool_pre_ping=True)
_TestSession = async_sessionmaker(_test_engine, class_=AsyncSession, expire_on_commit=False)

FIXTURES_DIR = Path(__file__).parent / "fixtures"
FIXTURES_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Low-level DB helper
# ---------------------------------------------------------------------------


async def _get_session() -> AsyncSession:
    return _TestSession()


# ---------------------------------------------------------------------------
# HTTP client fixture
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def client() -> AsyncClient:
    """Fresh HTTPX async client per test function (maintains cookie jar)."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac


# ---------------------------------------------------------------------------
# Helper: do a full login+2FA flow and return access_token string
# ---------------------------------------------------------------------------


async def _full_login(username: str, password: str, totp_secret: str) -> str:
    """Perform full login flow and return access_token."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        resp = await ac.post(
            "/api/auth/login",
            json={"username": username, "password": password},
        )
        assert resp.status_code == 200, f"Login failed: {resp.text}"

        otp_code = pyotp.TOTP(totp_secret).now()
        resp2 = await ac.post("/api/auth/2fa/verify", json={"code": otp_code})
        assert resp2.status_code == 200, f"2FA verify failed: {resp2.text}"
        return resp2.json()["access_token"]


# ---------------------------------------------------------------------------
# Session-scoped admin user + token
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture(scope="session")
async def admin_user() -> dict:
    """
    Create a QA admin user at session start.
    Returns dict with id, username, password, totp_secret.
    Cleaned up at session end.
    """
    uid_short = uuid.uuid4().hex[:8]
    username = f"qa_admin_{uid_short}"
    password = "QaAdmin123!"
    totp_secret = pyotp.random_base32()

    async with _TestSession() as session:
        user = User(
            username=username,
            password_hash=hash_password(password),
            role="admin",
            full_name=f"QA Admin {uid_short}",
            totp_secret=totp_secret,
            is_active=True,
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        user_id = user.id

    yield {"id": user_id, "username": username, "password": password, "totp_secret": totp_secret}

    async with _TestSession() as session:
        await session.execute(delete(ImportHistory).where(ImportHistory.uploaded_by == user_id))
        await session.execute(delete(User).where(User.id == user_id))
        await session.commit()


@pytest_asyncio.fixture(scope="session")
async def admin_token(admin_user: dict) -> str:
    """Full login+2FA flow for the session-scoped admin. Returns access_token."""
    return await _full_login(
        admin_user["username"], admin_user["password"], admin_user["totp_secret"]
    )


@pytest.fixture(scope="session")
def admin_headers(admin_token: str) -> dict:
    """Authorization header dict for the session admin."""
    return {"Authorization": f"Bearer {admin_token}"}


# ---------------------------------------------------------------------------
# Function-scoped helper users
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def employee_user() -> dict:
    """
    Create a QA employee user for a single test.
    Returns dict with id, username, password, totp_secret.
    """
    uid_short = uuid.uuid4().hex[:8]
    username = f"qa_emp_{uid_short}"
    password = "QaEmp123!"
    totp_secret = pyotp.random_base32()

    async with _TestSession() as session:
        user = User(
            username=username,
            password_hash=hash_password(password),
            role="employee",
            full_name=f"QA Employee {uid_short}",
            totp_secret=totp_secret,
            is_active=True,
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        user_id = user.id

    yield {"id": user_id, "username": username, "password": password, "totp_secret": totp_secret}

    async with _TestSession() as session:
        await session.execute(
            delete(AttendanceLog).where(AttendanceLog.employee_id == user_id)
        )
        await session.execute(delete(User).where(User.id == user_id))
        await session.commit()


@pytest_asyncio.fixture
async def employee_token(employee_user: dict) -> str:
    """access_token for the function-scoped employee user."""
    return await _full_login(
        employee_user["username"], employee_user["password"], employee_user["totp_secret"]
    )


@pytest_asyncio.fixture
async def manager_user() -> dict:
    """Create a QA manager user for a single test."""
    uid_short = uuid.uuid4().hex[:8]
    username = f"qa_mgr_{uid_short}"
    password = "QaMgr123!"
    totp_secret = pyotp.random_base32()

    async with _TestSession() as session:
        user = User(
            username=username,
            password_hash=hash_password(password),
            role="manager",
            full_name=f"QA Manager {uid_short}",
            totp_secret=totp_secret,
            is_active=True,
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        user_id = user.id

    yield {"id": user_id, "username": username, "password": password, "totp_secret": totp_secret}

    async with _TestSession() as session:
        await session.execute(delete(User).where(User.id == user_id))
        await session.commit()


@pytest_asyncio.fixture
async def manager_token(manager_user: dict) -> str:
    """access_token for the function-scoped manager user."""
    return await _full_login(
        manager_user["username"], manager_user["password"], manager_user["totp_secret"]
    )


@pytest_asyncio.fixture
async def fresh_user_no_2fa() -> dict:
    """Create a user WITHOUT totp_secret (simulates first login)."""
    uid_short = uuid.uuid4().hex[:8]
    username = f"qa_new_{uid_short}"
    password = "QaNew123!"

    async with _TestSession() as session:
        user = User(
            username=username,
            password_hash=hash_password(password),
            role="employee",
            full_name=f"QA New {uid_short}",
            totp_secret=None,
            is_active=True,
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        user_id = user.id

    yield {"id": user_id, "username": username, "password": password}

    async with _TestSession() as session:
        await session.execute(delete(User).where(User.id == user_id))
        await session.commit()


@pytest_asyncio.fixture
async def inactive_user() -> dict:
    """Create a QA user with is_active=False."""
    uid_short = uuid.uuid4().hex[:8]
    username = f"qa_inactive_{uid_short}"
    password = "QaInactive123!"

    async with _TestSession() as session:
        user = User(
            username=username,
            password_hash=hash_password(password),
            role="employee",
            full_name=f"QA Inactive {uid_short}",
            totp_secret=None,
            is_active=False,
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        user_id = user.id

    yield {"id": user_id, "username": username, "password": password}

    async with _TestSession() as session:
        await session.execute(delete(User).where(User.id == user_id))
        await session.commit()


# ---------------------------------------------------------------------------
# Sample Excel file fixture
# ---------------------------------------------------------------------------


def _build_test_excel(path: Path) -> None:
    """
    Build a test Excel file with:
    - 25 valid attendance records (3 unique employees, dates in 2026-01)
    - 3 records with bad dates (graceful degradation test)
    - 2 records with fuzzy name variants (extra spaces)
    - 5 duplicate records (for dedup test)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["ФИО", "Время", "Событие", "Точка"]
    ws.append(headers)

    # Valid records: 3 employees × ~8 records each
    employees = [
        "Иванов Иван Иванович",
        "Петров Пётр Петрович",
        "Сидорова Анна Сергеевна",
    ]
    checkpoints = ["Вход А", "Вход Б", "Турникет 1"]
    event_types = ["вход", "выход"]

    row_count = 0
    # Use Jan 13 (day=13 > 12: unambiguous with dayfirst=True), Mon 2026-01-13
    base_dt = datetime(2026, 1, 13, 8, 0, 0)

    for emp in employees:
        for day_offset in range(5):  # Mon–Fri
            for hour_offset, evt in [(0, "вход"), (9, "выход")]:
                dt = base_dt.replace(
                    day=base_dt.day + day_offset,
                    hour=8 + hour_offset,
                    minute=row_count % 30,
                )
                cp = checkpoints[row_count % len(checkpoints)]
                ws.append([emp, dt.strftime("%Y-%m-%d %H:%M:%S"), evt, cp])
                row_count += 1

    # Rows with bad dates (graceful degradation)
    ws.append(["Иванов Иван Иванович", "not-a-date", "вход", "Вход А"])
    ws.append(["Петров Пётр Петрович", "999-99-99 99:99", "выход", "Вход Б"])
    ws.append(["Сидорова Анна Сергеевна", "invalid-date-string", "вход", "Турникет 1"])

    # Fuzzy name variants (extra/double spaces) — same employee
    # Use day=20 (> 12, unambiguous)
    dt_fuzzy = datetime(2026, 1, 20, 9, 0, 0)
    ws.append(["Иванов Иван Иванович ", dt_fuzzy.strftime("%Y-%m-%d %H:%M:%S"), "вход", "Вход А"])
    ws.append(["Иванов Иван  Иванович", dt_fuzzy.strftime("%Y-%m-%d %H:%M:%S"), "выход", "Вход А"])

    # 5 duplicate records (repeat first 5 valid rows)
    for row_data in list(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
        ws.append(list(row_data))

    wb.save(str(path))


@pytest.fixture(scope="session")
def sample_excel_path() -> Path:
    """Path to the test Excel file. Generated once per session."""
    path = FIXTURES_DIR / "test_data.xlsx"
    _build_test_excel(path)
    yield path


@pytest.fixture(scope="session")
def invalid_excel_path() -> Path:
    """Path to an Excel file with ONLY invalid date rows."""
    path = FIXTURES_DIR / "invalid_dates.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ФИО", "Время", "Событие", "Точка"])
    ws.append(["Тестов Тест", "not-a-date", "вход", "Вход А"])
    ws.append(["Тестов Тест", "bad-date-string", "выход", "Вход А"])
    ws.append(["Тестов Тест", "99-99-9999", "вход", "Вход А"])
    wb.save(str(path))
    yield path


# ---------------------------------------------------------------------------
# DB helper exposed for tests that need to query the DB directly
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def db() -> AsyncSession:
    """Provides a raw DB session for direct DB queries in tests."""
    async with _TestSession() as session:
        yield session
